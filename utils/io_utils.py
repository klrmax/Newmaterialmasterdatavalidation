import os
import re
import pandas as pd
import openpyxl
import shutil
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Any, Callable, Set, Tuple

def load_text_file_to_set(filename: str) -> Set[str]:
    """Loads a .txt file (one item per line) into a lowercase set for fast lookups."""
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            # Use a set for O(1) average time complexity checks (very fast)
            # .strip() removes whitespace, .lower() ensures case-insensitivity
            return {line.strip().lower() for line in f if line.strip()}
    except FileNotFoundError:
        print(f" WARNING: Resource file '{filename}' not found. This check will be skipped.")
        return set()

def load_data(file_path: str) -> pd.DataFrame:
    """Loads data from an Excel file into a pandas DataFrame."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file was not found at: {file_path}")
    print(f"Loading data from '{os.path.basename(file_path)}'...")
    return pd.read_excel(file_path)

def save_data_to_excel(file_path: str, results_df: pd.DataFrame):
    """
    Opens an Excel workbook and updates multiple columns with new data,
    preserving all existing formatting.
    """
    print(f"Opening workbook to write results...")
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        raise FileNotFoundError(f"Failed to open workbook. Not found at: {file_path}")

    # Create a mapping from column name to its numerical index
    header = {cell.value: cell.column for cell in sheet[1]}
    
    # Update each column specified in the results DataFrame
    for col_name in results_df.columns:
        if col_name not in header:
            print(f"⚠️ WARNING: Target column '{col_name}' not found in Excel. It will be skipped.")
            continue
        
        target_col_num = header[col_name]
        print(f"Writing data to column '{col_name}'...")
        for index, value in results_df[col_name].items():
            sheet.cell(row=index + 2, column=target_col_num, value=value)
            
    try:
        workbook.save(file_path)
    except PermissionError:
        raise PermissionError(f"Could not save. Is the file '{os.path.basename(file_path)}' open in Excel?")
    
def save_data_to_new_excel(input_file_path: str, output_file_path: str, results_df: pd.DataFrame):
    """
    Duplicates an Excel template file and updates columns in the new file
    with data from a DataFrame, preserving all original formatting.

    Args:
        input_file_path (str): The path to the source Excel template file.
        output_file_path (str): The path where the new output Excel file will be saved.
        results_df (pd.DataFrame): The DataFrame containing the data to write.
                                   The DataFrame's index should correspond to the Excel rows
                                   (e.g., index 0 writes to row 2).
    """
    # --- 1. Duplicate the template file ---
    print(f"Duplicating template from '{input_file_path}' to '{output_file_path}'...")
    try:
        shutil.copy(input_file_path, output_file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"Failed to find the template file at: {input_file_path}")
    except Exception as e:
        raise IOError(f"Failed to create a copy of the template. Error: {e}")

    # --- 2. Open the NEW workbook to write the results ---
    print("Opening the new workbook to write results...")
    try:
        workbook = openpyxl.load_workbook(output_file_path)
        sheet = workbook.active
    except FileNotFoundError:
        # This is unlikely to happen since we just created it, but good for safety
        raise FileNotFoundError(f"Failed to open the new workbook at: {output_file_path}")

    # --- 3. Map header names to column numbers for easy lookup ---
    # We read the header from the first row (row 1)
    header = {cell.value: cell.column for cell in sheet[1]}
    
    # --- 4. Write data from the DataFrame to the corresponding columns ---
    for col_name in results_df.columns:
        if col_name not in header:
            print(f"⚠️ WARNING: Target column '{col_name}' not found in the Excel template. It will be skipped.")
            continue
        
        target_col_num = header[col_name]
        print(f"Writing data to column '{col_name}'...")
        
        # The DataFrame index corresponds to the data rows.
        # We add 2 because Excel is 1-indexed and we skip the header row.
        # (DataFrame index 0 -> Excel row 2)
        for index, value in results_df[col_name].items():
            # Check if the value is a pandas NaN, and convert it to None for openpyxl
            if pd.isna(value):
                value = None 
            sheet.cell(row=index + 2, column=target_col_num, value=value)
            
    # --- 5. Save the updated workbook ---
    print(f"Saving final workbook to '{output_file_path}'...")
    try:
        workbook.save(output_file_path)
        print("✅ Successfully saved the file!")
    except PermissionError:
        # This is less likely now but still possible in some scenarios
        raise PermissionError(f"Could not save. Is the new file '{os.path.basename(output_file_path)}' open?")
    except Exception as e:
        raise IOError(f"An unexpected error occurred while saving the file: {e}")

